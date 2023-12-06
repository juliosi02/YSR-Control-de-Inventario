import sys
from PyQt5.QtWidgets import QApplication, QMainWindow, QVBoxLayout, QPushButton, QWidget, QFileDialog
import sqlite3
import openpyxl

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.init_ui()

    def init_ui(self):
        self.setWindowTitle('Gestor de Inventario')

        # Botones de exportar e importar
        export_button = QPushButton('Exportar Base de Datos a Excel', self)
        export_button.clicked.connect(self.exportar_a_excel)

        import_button = QPushButton('Importar desde Excel a Base de Datos', self)
        import_button.clicked.connect(self.importar_desde_excel)

        # Configurar el diseño de la ventana
        layout = QVBoxLayout()
        layout.addWidget(export_button)
        layout.addWidget(import_button)

        central_widget = QWidget()
        central_widget.setLayout(layout)
        self.setCentralWidget(central_widget)

    def exportar_a_excel(self):
        # Conectar a la base de datos SQLite
        conn = sqlite3.connect('./database/db.db')
        cursor = conn.cursor()

        # Obtener datos de la base de datos
        cursor.execute("SELECT * FROM HerramientasManuales")
        datos = cursor.fetchall()

        # Crear un nuevo archivo Excel
        libro_excel = openpyxl.Workbook()
        hoja_excel = libro_excel.active
        columnas = [description[0] for description in cursor.description]

    # Escribir las cabeceras en el archivo Excel
        for columna, nombre_columna in enumerate(columnas, start=1):
            hoja_excel.cell(row=1, column=columna, value=nombre_columna)

        # Escribir datos en el archivo Excel
        for fila, dato in enumerate(datos, start=2):
            for columna, valor in enumerate(dato, start=1):
                hoja_excel.cell(row=fila, column=columna, value=valor)

        # Guardar el archivo Excel con el filtro específico
        filtro = 'Archivos Excel (*.xlsx);;Todos los archivos (*)'
        archivo_excel, _ = QFileDialog.getSaveFileName(self, 'Guardar como', '', filtro)

        if archivo_excel:
            if not archivo_excel.endswith('.xlsx'):
                archivo_excel += '.xlsx'
    
        libro_excel.save(archivo_excel)

        conn.close()
    def importar_desde_excel(self):
         # Seleccionar el archivo Excel
        archivo_excel = QFileDialog.getOpenFileName(self, 'Seleccionar archivo Excel', '', 'Archivos Excel (*.xlsx)')[0]

        if archivo_excel:
            # Conectar a la base de datos SQLite
            conn = sqlite3.connect('./database/databaseprueba.db')
            cursor = conn.cursor()

            # Abrir el archivo Excel
            libro_excel = openpyxl.load_workbook(archivo_excel)
            hoja_excel = libro_excel.active

            # Obtener los nombres de las columnas desde la primera fila del archivo Excel
            columnas = [cell.value for cell in hoja_excel[1]]

            # Recorrer las filas del archivo Excel e insertar los datos en la base de datos
            for fila_excel in hoja_excel.iter_rows(min_row=2, values_only=True):
                # Crear un diccionario que asocie el nombre de la columna con el valor de la fila
                datos = dict(zip(columnas, fila_excel))

                # Insertar los datos en la base de datos
                cursor.execute("INSERT INTO HerramientasManuales ({}) VALUES ({})".format(
                    ', '.join(datos.keys()),
                    ', '.join('?' * len(datos))
                ), tuple(datos.values()))

            conn.commit()
            conn.close()

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())